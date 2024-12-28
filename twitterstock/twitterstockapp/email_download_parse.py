import imaplib
import email
from email.header import decode_header
import sys, os
from openpyxl import load_workbook
import zipfile

sys.path.append(os.path.dirname(os.path.abspath(os.path.dirname(__file__))))
from db.postgres_orm import PostgresOrm


postgresOrm = PostgresOrm()
''' 
뱅크샐러드 앱에서 이메일 데이터 내보내기 하고
 1. 메일에서 파일 다운로드
 2. 파일 암호 풀고 데이터 읽기.
 3. postgresql에다가 저장(삭제 후 인서트로)
'''

EMAIL = "yangsungsim227@gmail.com"
PASSWORD = "vche kvrq dlql dxcz"
IMAP_SERVER = "imap.gmail.com"

DOWNLOAD_FOLDER = "C:/Temp/excel"
dir = "C:/Temp/excel"
extract_to = "C:/Temp/excel/extract"

column_to_search = 2
os.makedirs(DOWNLOAD_FOLDER, exist_ok=True)

def clean_subject(subject):
    decoded = decode_header(subject)
    subject = ''
    for part, encoding in decoded:
        if isinstance(part, bytes):
            subject += part.decode(encoding or 'utf-8')
        else:
            subject += part
    return subject

def download_attachments(msg, download_folder):
    for part in msg.walk():
        if part.get_content_disposition() == "attachment":
            filename = part.get_filename()
            if filename:
                filepath = os.path.join(download_folder, filename)
                with open(filepath, "wb") as f:
                    f.write(part.get_payload(decode=True))
                print(f"첨부 파일 저장 완료: {filepath}")
                return filepath
    return None

def extract_zipfile():
    entries = os.listdir(dir)
    if len(entries) == 0:
        print("nothing here")
    else:
        with zipfile.ZipFile(dir+"/"+entries[0], 'r') as zip_ref:
            zip_ref.setpassword('1234'.encode())
            zip_ref.extractall(extract_to)


def search_value_from_sheet(sheet, search_keyword, found_cell):
    for cell in sheet.iter_cols(min_col=column_to_search, max_col=column_to_search, values_only=False):
        for c in cell:
            if c.value == search_keyword:
                found_cell = c
                break
        if found_cell:
            break
    return found_cell

def search_value_from_certain_row(sheet, search_keyword, from_row, found_cell):
    for cell in sheet.iter_cols(min_col=column_to_search, max_col=column_to_search, min_row=from_row.row, values_only=False):
        for c in cell:
            if c.value == search_keyword:
                found_cell = c
                break
        if found_cell:
            break
    return found_cell


def parse_excel():
    entries = os.listdir(extract_to)
    wb = load_workbook(extract_to+"/"+entries[0])
    sheet = wb.active

    found_cell1 = None
    search_value = "3.재무현황"
    found_cell1 = search_value_from_sheet(sheet=sheet, search_keyword=search_value, found_cell=found_cell1)

    search_value = "총자산"
    found_cell2 = None
    found_cell2 = search_value_from_certain_row(sheet=sheet, search_keyword=search_value, from_row=found_cell1, found_cell=found_cell2)

    search_value = "5.투자현황"
    found_cell3 = None
    found_cell3 = search_value_from_sheet(sheet=sheet, search_keyword=search_value, found_cell=found_cell3)

    search_value = "총계"
    found_cell4 = None
    found_cell4 = search_value_from_certain_row(sheet=sheet, search_keyword=search_value, from_row=found_cell3, found_cell=found_cell3)

    search_value = "월수입 총계"
    found_cell6 = None
    found_cell6 = search_value_from_sheet(sheet=sheet, search_keyword=search_value, found_cell=found_cell6)

    search_value = "월지출 총계"
    found_cell7 = None
    found_cell7 = search_value_from_sheet(sheet=sheet, search_keyword=search_value, found_cell=found_cell7)


    # 재무현황
    kind = None
    name = None
    amount = None
    asset_data = []
    for col in sheet.iter_cols(min_col=column_to_search, max_col=column_to_search, min_row=found_cell1.row+4, max_row=found_cell2.row-1, values_only=False):
        for c in col:
            if c.value != None and len(c.value) > 0:
                kind = c.value
            name = sheet.cell(row=c.row, column=c.column+1).value
            amount = sheet.cell(row=c.row, column=c.column+3).value
            if name is None or amount is None or amount == 0:
                continue
            asset_data.append([kind, name, amount])

    # 투자현황
    kind = None
    name = None
    prev_amount = None
    cur_amount = None
    profit_rate = None
    investment_data = []
    for col in sheet.iter_cols(min_col=column_to_search, max_col=column_to_search, min_row=found_cell3.row+3, max_row=found_cell4.row-1, values_only=False):
        for c in col:
            kind = sheet.cell(row=c.row, column=c.column+1).value
            name = sheet.cell(row=c.row, column=c.column+2).value
            prev_amount = sheet.cell(row=c.row, column=c.column+4).value
            cur_amount =  sheet.cell(row=c.row, column=c.column+5).value
            profit_rate = sheet.cell(row=c.row, column=c.column+6).value
            investment_data.append([kind, name, prev_amount, cur_amount, profit_rate])


    # 현금흐름현황
    name = None
    amount = None
    date = None
    income_data = []
    for col in sheet.iter_cols(min_col=column_to_search, max_col=column_to_search, min_row=found_cell6.row, max_row=found_cell6.row, values_only=False):
        for c in col:
            name = c.value
            for i in range(4, 16):
                sum_amount = 0
                for r in range(1, 6):
                    sum_amount += sheet.cell(row=c.row-r, column=c.column+i).value
                amount = sum_amount
                date = sheet.cell(row=c.row-6, column=c.column+i).value
                income_data.append([name, amount, date])

    name = None
    amount = None
    date = None
    spend_data = []
    for col in sheet.iter_cols(min_col=column_to_search, max_col=column_to_search, min_row=found_cell7.row, max_row=found_cell7.row, values_only=False):
        for c in col:
            name = c.value
            for i in range(4, 16):
                sum_amount = 0
                for r in range(1, 18):
                    sum_amount += sheet.cell(row=c.row-r, column=c.column+i).value
                amount = sum_amount
                date = sheet.cell(row=c.row-25, column=c.column+i).value
                spend_data.append([name, amount, date])
    
    return asset_data, investment_data, income_data, spend_data
                

def save_into_table(asset_data, investment_data, income_data, spend_data):
    postgresOrm.insert("asset", ["name", "kind", "amount"], asset_data)
    postgresOrm.insert("investment", ["kind", "name", "prev_value", "cur_value", "profit_rate"], investment_data)
    postgresOrm.insert("cashflow", ["name", "amount", "date"], income_data)
    postgresOrm.insert("cashflow",  ["name", "amount", "date"], spend_data)
    postgresOrm.close()
    print("data insert finish")

def main():
    try:
        # Gmail 연결
        mail = imaplib.IMAP4_SSL(IMAP_SERVER)
        mail.login(EMAIL, PASSWORD)
        mail.select("inbox")

        # 이메일 검색 (첨부 파일 포함된 최신 이메일)
        status, messages = mail.search(None, 'ALL FROM "export-noreply@banksalad.com"')
        email_ids = messages[0].split()

        for email_id in reversed(email_ids):
            res, msg_data = mail.fetch(email_id, "(RFC822)")
            for response_part in msg_data:
                if isinstance(response_part, tuple):
                    msg = email.message_from_bytes(response_part[1])
                    subject = clean_subject(msg["Subject"])
                    print(f"이메일 제목: {subject}")

                    #첨부파일 다운로드
                    filepath = download_attachments(msg, DOWNLOAD_FOLDER)
                    if filepath:
                        #extract_zipfile()
                        asset_data, investment_data, income_data, spend_data = parse_excel()
                        save_into_table(asset_data, investment_data, income_data, spend_data)
    except Exception as e:
        print(f"오류 발생: {e}")
    finally:
        mail.logout()

if __name__ == "__main__":
    main()